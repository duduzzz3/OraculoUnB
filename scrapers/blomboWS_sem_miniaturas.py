# importações
from playwright.sync_api import sync_playwright
import pyautogui
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
import requests
import os
import re


# padroniza dimensões no formato A x B cm ou A x B x C cm

def padronizar_dimensoes(texto):
    if not texto or not str(texto).strip():
        return "Sem informação"

    numeros = re.findall(r"\d+(?:[.,]\d+)?", str(texto))
    if len(numeros) >= 3:
        return f"{numeros[0]} x {numeros[1]} x {numeros[2]} cm"
    if len(numeros) >= 2:
        return f"{numeros[0]} x {numeros[1]} cm"
    return "Sem informação"


# exporta dataframe para excel com miniaturas

def exportar_excel(df, arquivo_tmp, arquivo_final, mensagem):
    ordem = [
        "Título", "Autor", "Ano", "Técnica", "Dimensões",
        "Preço", "Descrição", "Link da obra", "Link da imagem da obra"
    ]

    for col in ordem:
        if col not in df.columns:
            df[col] = ""

    df = df.reindex(columns=ordem)

    with pd.ExcelWriter(arquivo_tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1, sheet_name="Dados")
        ws = writer.sheets["Dados"]

        ws["A1"] = mensagem

        header_row = 2
        data_start = 3
        headers_dict = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}

        if "Preço" in headers_dict:
            col_preco = get_column_letter(headers_dict["Preço"])
            for r in range(data_start, ws.max_row + 1):
                ws[f"{col_preco}{r}"].number_format = "0.00"

        ultima_col = get_column_letter(ws.max_column)
        ws.merge_cells(f"A1:{ultima_col}1")

    if os.path.exists(arquivo_final):
        os.remove(arquivo_final)
    os.rename(arquivo_tmp, arquivo_final)


with sync_playwright() as playwright:
    # abre navegador e páginas principal/detalhe
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    detail_page = context.new_page()

    # abre listagem inicial
    page.goto("https://blombo.com/obras/pinturas")

    dicionario_paginas = {}
    count = 1
    max_links = 150
    links_vistos = set()
    count_links = 0
    parar_tudo = False

    # percorre as páginas da listagem
    while True:
        proxima_pagina = page.locator("div.pages a.next.i-next[title='Próximo']").first
        artes = page.locator("div[data-product-id]")
        total = artes.count()
        print(f"{count}ª Página -> Obras encontradas: {total}")

        if total == 0:
            print("Nenhuma obra encontrada nessa página. Encerrando.")
            break

        lista_dicionarios = []

        # percorre os cards da página
        for i in range(total):
            arte = artes.nth(i)

            link_arte = arte.locator("a.product-image").get_attribute("href") or "Sem informação"
            titulo_arte = arte.locator(".product__name").inner_text().strip() or "Sem informação"
            autor = arte.locator(".product__artist").inner_text().strip() or "Sem informação"
            link_imagem = arte.locator("img").get_attribute("src") or "Sem informação"
            tecnica = arte.locator(".product__technique").inner_text().strip() or "Sem informação"
            dimensoes = padronizar_dimensoes(arte.locator(".product__dimensions").inner_text().strip())

            ano = arte.locator("span.product__technique + span").inner_text().strip()
            ano = ano.lstrip(", ").strip() if ano else "Sem informação"

            if link_arte in links_vistos:
                continue

            links_vistos.add(link_arte)
            count_links += 1

            if count_links > max_links:
                parar_tudo = True
                break

            preco = "Sem informação"
            preco_loc = arte.locator("p.special-price span.price").first
            if preco_loc.count() == 0:
                preco_loc = arte.locator("span.price[id^='product-price-']").first
            if preco_loc.count() == 0:
                preco_loc = arte.locator("span.price:not([id^='old-price-'])").first
            if preco_loc.count() > 0:
                preco = float(preco_loc.inner_text().strip().replace("R$", "").strip().replace(".", "").replace(",", "."))

            descricao = "Sem informação"

            # abre página interna para buscar descrição
            if link_arte != "Sem informação":
                detail_page.goto(link_arte, wait_until="domcontentloaded")
                desc_loc = detail_page.locator("p.c-product__artist").first
                if desc_loc.count() > 0:
                    descricao_texto = desc_loc.inner_text().strip()
                    if descricao_texto:
                        descricao = descricao_texto

            lista_dicionarios.append({
                "Título": titulo_arte,
                "Autor": autor,
                "Ano": ano,
                "Técnica": tecnica,
                "Dimensões": dimensoes,
                "Preço": preco,
                "Descrição": descricao,
                "Link da obra": link_arte,
                "Link da imagem da obra": link_imagem,
            })

            if count_links == max_links:
                print(f"Parando: cheguei em {max_links} links.")
                parar_tudo = True
                break

        dicionario_paginas[f"{count}ª Página"] = lista_dicionarios

        if parar_tudo:
            break

        if proxima_pagina.count() == 0:
            print("Chegamos ao fim!")
            break

        url_atual = page.url
        proxima_pagina.click()
        page.wait_for_url(lambda url: url != url_atual)
        count += 1

    pyautogui.alert("Sistema Executado")
    context.close()
    browser.close()


# junta todas as páginas em uma lista única
lista_unificada = []
for lista in dicionario_paginas.values():
    lista_unificada.extend(lista)

df = pd.DataFrame(lista_unificada)
if not df.empty and "Link da obra" in df.columns:
    df = df.drop_duplicates(subset=["Link da obra"], keep="first")

site = "Blombo"
data_execucao = datetime.now().strftime("%d/%m/%Y")
mensagem = f"Planilha alimentada com dados do site ({site}) no dia {data_execucao}"

exportar_excel(df, "artes_BLOMBO_tmp.xlsx", "artes_BLOMBO.xlsx", mensagem)
