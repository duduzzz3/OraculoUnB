# -*- coding: utf-8 -*-
"""
Scraper LeilõesBR - sem miniaturas

Coleta lotes/obras a partir da busca da LeilõesBR e salva Excel compatível
com o padrão do projeto Oráculo Cultural.

Uso:
    python leiloesbrWS_sem_miniaturas.py --max-obras 100 --output artes_LeiloesBR.xlsx --headless true

Dependências:
    pip install playwright pandas openpyxl beautifulsoup4 lxml
    python -m playwright install chromium
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Any
from urllib.parse import parse_qsl, urlencode, urljoin, urlsplit, urlunsplit

import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


BASE_URL = "https://leiloesbr.com.br/"
START_URL = (
    "https://leiloesbr.com.br/busca_andamento.asp?"
    "pesquisa=&op=2&v=126&tp=|50696E74757261732065204772617675726173|&b=0&pag=2"
)
SEM_INFO = "Sem informação"

EXCEL_COLUMNS = [
    "Nome da obra",
    "Autor",
    "Preço",
    "Técnica",
    "Técnica original",
    "Dimensões",
    "Ano da obra",
    "Descrição",
    "Link da obra",
    "Link da imagem da obra",
]

BAD_LINK_PARTS = (
    "javascript:",
    "mailto:",
    "tel:",
    "facebook",
    "instagram",
    "twitter",
    "pinterest",
    "youtube",
    "tiktok",
    "whatsapp",
    "login",
    "cadastro",
    "contato",
    "ajuda",
    "blog",
)

BAD_IMAGE_PARTS = (
    "logo",
    "placeholder",
    "banner",
    "icon",
    "icone",
    "avatar",
    "sem-imagem",
    "sem_imagem",
    "default",
    "spacer",
    "blank",
    "loading",
    "favicon",
    "apple-touch",
)

GENERIC_AUTHOR_STARTS = (
    "pintura", "gravura", "escultura", "desenho", "lote", "par", "conjunto",
    "fruteira", "vaso", "imagem", "prato", "moeda", "selos", "selo", "cards",
    "espetacular", "autor", "sem título", "óleo", "oleo", "aquarela",
)

TECH_PATTERNS: list[tuple[str, list[str]]] = [
    ("Serigrafia", [
        r"\bserigrafi[ao]s?\b", r"\bserigraph\b", r"\bscreen\s*print\b", r"\bscreenprint\b",
    ]),
    ("Xilogravura", [
        r"\bxilogravura\b", r"\bwoodcut\b", r"\blinocut\b", r"\blinogravura\b",
    ]),
    ("Gravura em metal", [
        r"\bgravura\s+em\s+metal\b", r"\bmetal\s+engraving\b", r"\bintaglio\b", r"\bengraving\b",
    ]),
    ("Litografia", [
        r"\blitografia\b", r"\blithograph(?:y)?\b",
    ]),
    ("Água-forte", [
        r"\b[aá]gua[-\s]?forte\b", r"\betching\b",
    ]),
    ("Água-tinta", [
        r"\b[aá]gua[-\s]?tinta\b", r"\baquatint\b",
    ]),
    ("Ponta-seca", [
        r"\bponta[-\s]?seca\b", r"\bdrypoint\b",
    ]),
    ("Pintura a óleo", [
        r"\b[oó]leo\s+s(?:obre|\.)?\s*(?:tela|juta|madeira|papel|eucatex|placa)\b",
        r"\bo\.?\s*s\.?\s*/?\s*(?:tela|juta|madeira|papel|eucatex|placa)\b",
        r"\bo\.?\s*s\.?\s*(?:tela|juta|madeira|papel|eucatex|placa)\b",
        r"\boil\s+on\s+canvas\b",
        r"\b[oó]leo\b",
    ]),
    ("Pintura acrílica", [
        r"\bacr[ií]lica\s+s(?:obre|\.)?\s*(?:tela|papel|madeira|eucatex|cart[aã]o)\b",
        r"\ba\.?\s*s\.?\s*/?\s*(?:tela|papel|madeira|eucatex|cart[aã]o)\b",
        r"\bacrylic\b", r"\bacr[ií]lica\b",
    ]),
    ("Aquarela", [
        r"\baquarela(?:do|da|s)?\b", r"\bwatercolor\b",
    ]),
    ("Guache", [
        r"\bguache\b", r"\bgouache\b",
    ]),
    ("Nanquim / tinta", [
        r"\bnanquim\b", r"\btinta\b", r"\bink\b",
    ]),
    ("Técnica mista", [
        r"\bt[ée]cnica\s+mista\b", r"\bt\.?\s*m\.?\s*s\.?\s*/?\s*(?:papel|tela|cart[aã]o)?\b",
        r"\bmixed\s+media\b", r"\bmista\s+s(?:obre|\.)?\s*(?:papel|tela|cart[aã]o)\b",
        r"\bcolagem\b", r"\bcollage\b",
    ]),
    ("Stencil", [
        r"\bstencil\b",
    ]),
    ("Giclée", [
        r"\bgicl[ée]e?\b", r"\bgiclee\b",
    ]),
    ("Desenho", [
        r"\bdesenho\b", r"\bdrawing\b", r"\bgrafite\b", r"\bcarv[aã]o\b",
    ]),
    ("Fotografia", [
        r"\bfotografia\b", r"\bphotograph(?:y)?\b",
    ]),
    ("Escultura", [
        r"\bescultura\b", r"\bsculpture\b", r"\bresina\s+acr[ií]lica\b",
    ]),
    ("Gravura", [
        r"\bgravura\b", r"\bprintmaking\b", r"\bprint\b",
    ]),
]

DIM_RE = re.compile(
    r"(?<!R\$)(?<!\$)\b(\d{1,4}(?:[,.]\d{1,2})?)\s*(?:cm)?\s*"
    r"(?:x|×|X|por)\s*"
    r"(\d{1,4}(?:[,.]\d{1,2})?)\s*(?:cm)?"
    r"(?:\s*(?:x|×|X|por)\s*(\d{1,4}(?:[,.]\d{1,2})?)\s*(?:cm)?)?",
    re.I,
)

PRICE_RE = re.compile(
    r"R\$\s*(\d{1,3}(?:[.\s]\d{3})*(?:,\d{2})?|\d+(?:,\d{2})?)",
    re.I,
)

YEAR_FULL_RE = re.compile(r"\b(18|19|20)\d{2}\b")
YEAR_NEAR_DATE_RE = re.compile(
    r"\b(?:datad[ao]s?|data|dated|assinado\s+e\s+datado|verso)\D{0,35}"
    r"((?:18|19|20)\d{2}|\d{2})\b",
    re.I,
)


@dataclass
class CardInfo:
    link: str
    titulo_bruto: str = SEM_INFO
    preco: float | str = SEM_INFO
    imagem: str = SEM_INFO
    leiloeiro: str = SEM_INFO


@dataclass
class Obra:
    nome: str = SEM_INFO
    autor: str = SEM_INFO
    preco: float | str = SEM_INFO
    tecnica: str = SEM_INFO
    tecnica_original: str = SEM_INFO
    dimensoes: str = SEM_INFO
    ano: str = SEM_INFO
    descricao: str = SEM_INFO
    link: str = SEM_INFO
    imagem: str = SEM_INFO


def limpar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    texto = str(valor)
    texto = texto.replace("\ufeff", "").replace("\xa0", " ").replace("&nbsp;", " ")
    texto = re.sub(r"[\t\r\f\v]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def sem_info(valor: Any) -> str:
    texto = limpar_texto(valor)
    return texto if texto else SEM_INFO


def str_to_bool(valor: Any) -> bool:
    if isinstance(valor, bool):
        return valor
    return str(valor).strip().lower() in {"1", "true", "sim", "yes", "y", "s"}


def url_absoluta(url: str | None, base: str = BASE_URL) -> str:
    if not url:
        return SEM_INFO
    url = limpar_texto(url)
    if not url or url.startswith("#"):
        return SEM_INFO
    if url.startswith("//"):
        return "https:" + url
    return urljoin(base, url)


def alterar_parametro_pag(url: str, pag: int) -> str:
    partes = urlsplit(url)
    query = dict(parse_qsl(partes.query, keep_blank_values=True))
    query["pag"] = str(pag)
    nova_query = urlencode(query, doseq=True, safe="|:/")
    return urlunsplit((partes.scheme, partes.netloc, partes.path, nova_query, partes.fragment))


def pegar_pag_inicial(url: str) -> int:
    try:
        query = dict(parse_qsl(urlsplit(url).query, keep_blank_values=True))
        return max(1, int(query.get("pag", "1")))
    except Exception:
        return 1


def link_lote_valido(url: str) -> bool:
    if not url or url == SEM_INFO:
        return False
    low = url.lower()
    if any(x in low for x in BAD_LINK_PARTS):
        return False
    return "abre_catalogo.asp" in low and "|" in url


def imagem_valida(url: str) -> bool:
    if not url or url == SEM_INFO:
        return False
    low = url.lower()
    if low.startswith("data:"):
        return False
    if any(x in low for x in BAD_IMAGE_PARTS):
        return False
    if not re.search(r"\.(jpe?g|png|webp|avif)(?:[?#].*)?$", low):
        # A LeilõesBR costuma usar JPG em CloudFront; se não houver extensão, rejeita.
        return False
    return True


def melhor_url_srcset(srcset: str | None) -> str:
    if not srcset:
        return ""
    partes = [p.strip().split(" ")[0] for p in srcset.split(",") if p.strip()]
    return partes[-1] if partes else ""


def normalizar_preco(texto: str) -> float | str:
    texto = limpar_texto(texto)
    m = PRICE_RE.search(texto)
    if not m:
        return SEM_INFO
    raw = m.group(1).replace(".", "").replace(" ", "").replace(",", ".")
    try:
        return round(float(raw), 2)
    except ValueError:
        return SEM_INFO


def formatar_numero_dimensao(valor: str) -> str:
    valor = valor.replace(",", ".")
    try:
        n = float(valor)
        if n.is_integer():
            return str(int(n))
        return f"{n:.2f}".rstrip("0").rstrip(".").replace(".", ",")
    except ValueError:
        return valor.replace(".", ",")


def normalizar_dimensoes(texto: str) -> str:
    texto = limpar_texto(texto)
    if not texto:
        return SEM_INFO

    # Remove preços para evitar confundir R$ 1.500,00 com dimensão.
    texto_sem_preco = PRICE_RE.sub(" ", texto)
    match = DIM_RE.search(texto_sem_preco)
    if not match:
        return SEM_INFO

    partes = [
        formatar_numero_dimensao(match.group(1)),
        formatar_numero_dimensao(match.group(2)),
    ]
    if match.group(3):
        partes.append(formatar_numero_dimensao(match.group(3)))
    return " x ".join(partes) + " cm"


def normalizar_ano(texto: str) -> str:
    texto = limpar_texto(texto)
    if not texto:
        return SEM_INFO

    texto_sem_biografia = re.sub(r"\((?:18|19|20)\d{2}\s*(?:[./\-–—]|\.{2})\s*(?:18|19|20)?\d{2}\)", " ", texto)
    texto_sem_biografia = re.sub(r"\b(?:18|19|20)\d{2}\s*/\s*(?:18|19|20)\d{2}\b", " ", texto_sem_biografia)

    m = YEAR_NEAR_DATE_RE.search(texto_sem_biografia)
    if m:
        raw = m.group(1)
        if len(raw) == 2:
            yy = int(raw)
            return str(2000 + yy) if yy <= 26 else str(1900 + yy)
        return raw

    # Fallback conservador: usa ano de obra apenas quando a frase sugere assinatura/datação.
    contexto = texto_sem_biografia.lower()
    if any(palavra in contexto for palavra in ("datad", "assinad", "ano", "dated")):
        m2 = YEAR_FULL_RE.search(texto_sem_biografia)
        if m2:
            return m2.group(0)

    return SEM_INFO


def normalizar_tecnica(texto: str) -> tuple[str, str]:
    texto = limpar_texto(texto)
    if not texto:
        return SEM_INFO, SEM_INFO

    # Prioridade: "escultura em resina acrílica" não deve virar "Pintura acrílica".
    m_escultura = re.search(r"\bescultura(?:\s+em\s+[A-Za-zÀ-ÿ\s]+?)?(?=,|\.|\s*\d|$)", texto, flags=re.I)
    if m_escultura:
        return "Escultura", limpar_texto(m_escultura.group(0))

    for tecnica_padrao, padroes in TECH_PATTERNS:
        for padrao in padroes:
            m = re.search(padrao, texto, flags=re.I)
            if m:
                return tecnica_padrao, limpar_texto(m.group(0))
    return SEM_INFO, SEM_INFO


def titulo_valido(titulo: str) -> bool:
    titulo = limpar_texto(titulo)
    if not titulo or titulo == SEM_INFO:
        return False
    if len(titulo) < 3:
        return False
    low = titulo.lower()
    ruins = (
        "login", "cadastre", "minha conta", "todas as peças", "leilões hoje",
        "novos", "blog", "ajuda", "contato", "facebook", "instagram",
    )
    return not any(r in low for r in ruins)


def parece_autor(parte: str) -> bool:
    parte = limpar_texto(parte).strip(" -:;")
    if not parte or len(parte) < 3 or len(parte) > 90:
        return False
    low = parte.lower()
    if any(low.startswith(g) for g in GENERIC_AUTHOR_STARTS):
        return False
    if normalizar_tecnica(parte)[0] != SEM_INFO:
        return False
    if normalizar_dimensoes(parte) != SEM_INFO:
        return False
    if re.search(r"\d{2,}", parte) and not re.search(r"\([^)]*(?:18|19|20)\d{2}[^)]*\)", parte):
        return False

    letras = re.sub(r"[^A-Za-zÀ-ÿ]", "", parte)
    if not letras:
        return False

    palavras = [p for p in re.split(r"\s+", re.sub(r"\([^)]*\)", "", parte).strip()) if p]
    if len(palavras) > 6:
        return False

    # Em muitos cards da LeilõesBR o autor vem em caixa alta antes da primeira vírgula.
    letras_maiusculas = sum(1 for c in letras if c.isupper())
    proporcao_maiusculas = letras_maiusculas / max(1, len(letras))
    if proporcao_maiusculas >= 0.55:
        return True

    # Também há autores em Title Case: "Durval Pereira, óleo sobre tela..."
    capitalizadas = 0
    avaliadas = 0
    for palavra in palavras:
        limpa = re.sub(r"[^A-Za-zÀ-ÿ]", "", palavra)
        if len(limpa) <= 1:
            continue
        avaliadas += 1
        if limpa[0].isupper():
            capitalizadas += 1
    return avaliadas > 0 and capitalizadas / avaliadas >= 0.75

def dividir_virgulas_fora_parenteses(texto: str) -> list[str]:
    partes: list[str] = []
    atual: list[str] = []
    nivel = 0
    for ch in texto:
        if ch == "(":
            nivel += 1
        elif ch == ")" and nivel > 0:
            nivel -= 1
        if ch == "," and nivel == 0:
            parte = limpar_texto("".join(atual))
            if parte:
                partes.append(parte)
            atual = []
        else:
            atual.append(ch)
    parte = limpar_texto("".join(atual))
    if parte:
        partes.append(parte)
    return partes



def limpar_nome_obra(parte: str, cortar_tecnica: bool = False) -> str:
    parte = limpar_texto(parte)
    parte = re.split(
        r"\b(?:assinado|assinada|datado|datada|numerado|numerada|medindo|dimensões|dimensao|moldura|sem moldura)\b",
        parte,
        maxsplit=1,
        flags=re.I,
    )[0]
    parte = re.split(r"\s+-\s*med\.?", parte, maxsplit=1, flags=re.I)[0]

    # Remove técnica e o que vem depois apenas quando já temos autor separado.
    # Sem autor, a técnica pode fazer parte do nome descritivo do lote ("Antiga gravura...").
    if cortar_tecnica:
        tecnica, tecnica_original = normalizar_tecnica(parte)
        if tecnica != SEM_INFO and tecnica_original != SEM_INFO:
            pos = parte.lower().find(tecnica_original.lower())
            prefixo = parte[:pos].strip(" ,.;:-")
            prefixo_low = prefixo.lower()
            prefixo_generico = prefixo_low in {"antiga", "antigo", "duas", "dois", "lote", "par", "conjunto"}
            if pos > 3 and len(prefixo) >= 3 and not prefixo_generico:
                parte = prefixo

    parte = parte.strip(" ,.;:-")
    parte = re.sub(r"\s+", " ", parte)
    return parte if parte else SEM_INFO


def extrair_autor_e_nome(texto: str) -> tuple[str, str]:
    texto = limpar_texto(texto)
    if not texto:
        return SEM_INFO, SEM_INFO

    texto = re.sub(r"\s+-\s*R\$\s*.*$", "", texto, flags=re.I).strip()
    partes = dividir_virgulas_fora_parenteses(texto)

    if len(partes) >= 2 and parece_autor(partes[0]):
        autor = partes[0].strip()

        # Caso: "Omar Pellegatta, óleo sobre juta, paisagem com igreja..."
        # A primeira parte após o autor pode ser técnica, e a obra vem depois.
        if len(partes) >= 3 and normalizar_tecnica(partes[1])[0] != SEM_INFO:
            fonte_nome = ", ".join(partes[2:])
        else:
            fonte_nome = ", ".join(partes[1:])

        nome = limpar_nome_obra(fonte_nome, cortar_tecnica=True)
        if not titulo_valido(nome):
            nome = limpar_nome_obra(partes[1], cortar_tecnica=False)
        return sem_info(autor), sem_info(nome)

    nome = limpar_nome_obra(texto, cortar_tecnica=False)
    if len(nome) > 180:
        # Evita salvar biografia/descrição inteira como nome.
        nome = re.split(r"[.;]", nome, maxsplit=1)[0].strip()
        if len(nome) > 180:
            nome = nome[:177].rstrip() + "..."
    return SEM_INFO, sem_info(nome)


def texto_de_tag(tag: Tag | None) -> str:
    if tag is None:
        return ""
    return limpar_texto(tag.get_text(" ", strip=True))


def atributo(tag: Tag | None, nome: str) -> str:
    if tag is None:
        return ""
    return limpar_texto(tag.get(nome, ""))


def extrair_texto_data_linktitle(card: Tag) -> str:
    link_lupa = card.select_one("a.smallmag[data-linktitle]")
    bruto = atributo(link_lupa, "data-linktitle")
    if not bruto:
        return ""
    try:
        soup = BeautifulSoup(bruto, "lxml")
        p = soup.find("p")
        if p:
            return texto_de_tag(p)
    except Exception:
        pass
    m = re.search(r"<p>(.*?)</p>", bruto, re.I | re.S)
    if m:
        return limpar_texto(BeautifulSoup(m.group(1), "lxml").get_text(" ", strip=True))
    return ""


def extrair_imagem_card(card: Tag) -> str:
    candidatos: list[str] = []

    img = card.select_one(".product-image img") or card.find("img")
    if img:
        candidatos.extend([
            atributo(img, "src"),
            atributo(img, "data-src"),
            atributo(img, "data-original"),
            atributo(img, "data-original-src"),
            melhor_url_srcset(atributo(img, "srcset")),
        ])

    link_lupa = card.select_one("a.smallmag[href]")
    if link_lupa:
        candidatos.append(atributo(link_lupa, "href"))

    for candidato in candidatos:
        url = url_absoluta(candidato)
        if imagem_valida(url):
            return url
    return SEM_INFO


def extrair_preco_card(card: Tag) -> float | str:
    seletores = [
        ".product-price.venda-price",
        ".venda-price",
        "[class*='price']",
        "a.stretched-link[title]",
    ]
    for seletor in seletores:
        for tag in card.select(seletor):
            texto = texto_de_tag(tag) or atributo(tag, "title")
            preco = normalizar_preco(texto)
            if preco != SEM_INFO:
                return preco
    return SEM_INFO


def extrair_titulo_card(card: Tag) -> str:
    candidatos = [
        atributo(card.select_one(".product-title a[title]"), "title"),
        atributo(card.select_one(".mostbidded__title a[title]"), "title"),
        atributo(card.select_one("a.stretched-link[title]"), "title"),
        atributo(card.select_one(".product-image img[alt]"), "alt"),
        extrair_texto_data_linktitle(card),
        texto_de_tag(card.select_one(".product-title h3")),
    ]
    for candidato in candidatos:
        candidato = limpar_texto(candidato)
        if titulo_valido(candidato):
            # Se o título da stretched-link vier com " - R$valor - leiloeiro", corta no preço.
            candidato = re.sub(r"\s+-\s*R\$\s*[\d.,]+.*$", "", candidato, flags=re.I).strip()
            return candidato
    return SEM_INFO


def extrair_leiloeiro_card(card: Tag) -> str:
    candidatos = [
        atributo(card.select_one(".mostbidded__info a[title]"), "title"),
        texto_de_tag(card.select_one(".mostbidded__info a")),
        atributo(card.select_one(".ellipsis-overflow a[title]"), "title"),
        texto_de_tag(card.select_one(".ellipsis-overflow a")),
    ]
    for c in candidatos:
        if c:
            return c
    return SEM_INFO


def extrair_links_obras_do_html(html: str, base_url: str = BASE_URL) -> list[CardInfo]:
    soup = BeautifulSoup(html, "lxml")
    cards = soup.select("div.mostbidded, div.oc-item")
    if not cards:
        cards = soup.select("div.product")

    encontrados: list[CardInfo] = []
    vistos: set[str] = set()

    for card in cards:
        link_tag = (
            card.select_one("a.stretched-link[href*='abre_catalogo.asp']")
            or card.select_one(".product-title a[href*='abre_catalogo.asp']")
            or card.select_one("a[href*='abre_catalogo.asp']")
        )
        href = atributo(link_tag, "href")
        link = url_absoluta(href, base_url)
        if not link_lote_valido(link) or link in vistos:
            continue

        titulo = extrair_titulo_card(card)
        if not titulo_valido(titulo):
            continue

        vistos.add(link)
        encontrados.append(
            CardInfo(
                link=link,
                titulo_bruto=titulo,
                preco=extrair_preco_card(card),
                imagem=extrair_imagem_card(card),
                leiloeiro=extrair_leiloeiro_card(card),
            )
        )

    return encontrados


def aceitar_cookies_e_popups(page: Page) -> None:
    seletores = [
        "button:has-text('Aceitar')",
        "button:has-text('Aceito')",
        "button:has-text('Accept')",
        "button:has-text('OK')",
        "button[aria-label*='fechar' i]",
        "button[aria-label*='close' i]",
        ".modal button.close",
        ".btn-close",
    ]
    for seletor in seletores:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0 and loc.is_visible(timeout=600):
                loc.click(timeout=1200)
                page.wait_for_timeout(300)
        except Exception:
            pass


def descobrir_cards_listagem(
    page: Page,
    start_url: str,
    max_obras: int = 0,
    max_pages: int = 200,
    delay: float = 0.4,
) -> list[CardInfo]:
    pagina_inicial = pegar_pag_inicial(start_url)
    todos: list[CardInfo] = []
    vistos: set[str] = set()
    sem_novos = 0

    for offset in range(max_pages):
        pagina = pagina_inicial + offset
        url = start_url if offset == 0 else alterar_parametro_pag(start_url, pagina)

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=45_000)
            aceitar_cookies_e_popups(page)
            page.wait_for_timeout(800)
            html = page.content()
        except PlaywrightTimeoutError:
            print(f"Página {pagina}: timeout ao carregar. Encerrando paginação.", file=sys.stderr)
            break
        except Exception as exc:
            print(f"Página {pagina}: erro ao carregar ({exc}). Encerrando paginação.", file=sys.stderr)
            break

        cards = extrair_links_obras_do_html(html, page.url)
        novos = [c for c in cards if c.link not in vistos]
        print(f"Página {pagina}: {len(cards)} links encontrados | {len(novos)} novos")

        if not cards:
            break

        if not novos:
            sem_novos += 1
            if sem_novos >= 2:
                print("Paginação encerrada: duas páginas seguidas sem links novos.")
                break
        else:
            sem_novos = 0

        for card in novos:
            vistos.add(card.link)
            todos.append(card)
            if max_obras > 0 and len(todos) >= max_obras:
                return todos

        if delay:
            time.sleep(delay)

    return todos


def meta_content(soup: BeautifulSoup, *selectors: str) -> str:
    for seletor in selectors:
        tag = soup.select_one(seletor)
        if tag:
            content = atributo(tag, "content")
            if content:
                return content
    return ""


def selecionar_descricao_detalhe(soup: BeautifulSoup) -> str:
    candidatos: list[str] = []

    meta_desc = meta_content(
        soup,
        "meta[property='og:description']",
        "meta[name='description']",
        "meta[name='twitter:description']",
    )
    if meta_desc:
        candidatos.append(meta_desc)

    for seletor in [
        ".lot-description",
        ".descricao-lote",
        ".descricao",
        ".product-description",
        "[class*='description']",
        "[class*='descricao']",
        "article",
        "main",
    ]:
        for tag in soup.select(seletor):
            texto = texto_de_tag(tag)
            if 20 <= len(texto) <= 1500:
                candidatos.append(texto)

    # Escolhe o texto mais informativo, sem virar página inteira/menu.
    candidatos = [c for c in candidatos if not texto_parece_menu(c)]
    if candidatos:
        candidatos.sort(key=len, reverse=True)
        return candidatos[0][:1200].strip()

    return ""


def texto_parece_menu(texto: str) -> bool:
    low = limpar_texto(texto).lower()
    if len(low) > 1600:
        return True
    sinais_menu = sum(
        1
        for termo in ["login", "cadastre", "minha conta", "categorias", "leilões hoje", "facebook", "instagram"]
        if termo in low
    )
    return sinais_menu >= 3


def extrair_imagem_detalhe(soup: BeautifulSoup, base_url: str, titulo: str = "") -> str:
    candidatos: list[tuple[int, str]] = []

    for seletor in [
        "meta[property='og:image']",
        "meta[property='og:image:secure_url']",
        "meta[name='twitter:image']",
    ]:
        url = meta_content(soup, seletor)
        url = url_absoluta(url, base_url)
        if imagem_valida(url):
            candidatos.append((20, url))

    titulo_low = titulo.lower()[:35]
    for img in soup.find_all("img"):
        urls = [
            atributo(img, "src"),
            atributo(img, "data-src"),
            atributo(img, "data-original"),
            atributo(img, "data-original-src"),
            melhor_url_srcset(atributo(img, "srcset")),
        ]
        alt = atributo(img, "alt").lower()
        width = atributo(img, "width")
        height = atributo(img, "height")
        score_base = 0
        try:
            score_base += min(int(re.sub(r"\D", "", width) or "0"), 2000) // 100
            score_base += min(int(re.sub(r"\D", "", height) or "0"), 2000) // 100
        except Exception:
            pass
        if titulo_low and titulo_low in alt:
            score_base += 10
        for raw in urls:
            url = url_absoluta(raw, base_url)
            if not imagem_valida(url):
                continue
            low = url.lower()
            score = score_base
            if "img_m" in low or "imagens" in low or "cloudfront" in low:
                score += 8
            candidatos.append((score, url))

    if not candidatos:
        return SEM_INFO
    candidatos.sort(key=lambda x: x[0], reverse=True)
    return candidatos[0][1]


def extrair_texto_label(soup: BeautifulSoup, labels: tuple[str, ...]) -> str:
    labels_low = tuple(l.lower() for l in labels)

    # Tenta linhas do tipo "Técnica: óleo sobre tela" ou "Dimensões 50 x 70 cm".
    linhas = [limpar_texto(x) for x in soup.get_text("\n", strip=True).splitlines() if limpar_texto(x)]
    for linha in linhas:
        low = linha.lower()
        for label in labels_low:
            if low.startswith(label):
                valor = re.sub(rf"^{re.escape(label)}\s*[:\-]?\s*", "", linha, flags=re.I)
                if valor and valor.lower() != label:
                    return valor

    # Tenta tabelas/listas com label e valor em células vizinhas.
    for tr in soup.select("tr"):
        cells = [texto_de_tag(c) for c in tr.find_all(["th", "td"])]
        if len(cells) >= 2:
            label = cells[0].lower()
            if any(l in label for l in labels_low):
                return cells[1]
    for li in soup.select("li"):
        texto = texto_de_tag(li)
        low = texto.lower()
        if any(low.startswith(l) for l in labels_low) and ":" in texto:
            return texto.split(":", 1)[1].strip()

    return ""


def montar_obra_a_partir_das_fontes(card: CardInfo, html_detalhe: str = "", base_url: str = BASE_URL) -> Obra:
    soup = BeautifulSoup(html_detalhe or "", "lxml") if html_detalhe else BeautifulSoup("", "lxml")

    titulo_detalhe = meta_content(soup, "meta[property='og:title']", "meta[name='twitter:title']")
    if not titulo_valido(titulo_detalhe):
        for seletor in ["h1", "h2", ".product-title", ".lot-title", "[class*='title']"]:
            titulo_detalhe = texto_de_tag(soup.select_one(seletor))
            if titulo_valido(titulo_detalhe) and not texto_parece_menu(titulo_detalhe):
                break

    # A listagem da LeilõesBR traz uma descrição bem limpa no title/alt do card.
    texto_base = card.titulo_bruto if titulo_valido(card.titulo_bruto) else titulo_detalhe
    descricao_detalhe = selecionar_descricao_detalhe(soup) if html_detalhe else ""
    descricao = texto_base
    if descricao_detalhe and len(descricao_detalhe) > len(descricao) and not texto_parece_menu(descricao_detalhe):
        descricao = descricao_detalhe

    autor, nome = extrair_autor_e_nome(texto_base)
    if nome == SEM_INFO and titulo_valido(titulo_detalhe):
        _, nome = extrair_autor_e_nome(titulo_detalhe)

    tecnica_label = extrair_texto_label(soup, ("técnica", "tecnica", "medium", "material", "suporte"))
    tecnica, tecnica_original = normalizar_tecnica(tecnica_label or descricao or texto_base)

    dimensoes_label = extrair_texto_label(soup, ("dimensões", "dimensoes", "medidas", "medida", "tamanho", "size"))
    dimensoes = normalizar_dimensoes(dimensoes_label or descricao or texto_base)

    ano_label = extrair_texto_label(soup, ("ano", "year", "data da obra"))
    ano = normalizar_ano(ano_label or descricao or texto_base)

    preco = card.preco
    if preco == SEM_INFO:
        preco_label = extrair_texto_label(soup, ("preço", "preco", "valor", "lance", "estimativa"))
        preco = normalizar_preco(preco_label or descricao or soup.get_text(" ", strip=True))

    imagem = card.imagem
    if not imagem_valida(imagem) and html_detalhe:
        imagem = extrair_imagem_detalhe(soup, base_url, nome if nome != SEM_INFO else texto_base)

    return Obra(
        nome=nome if titulo_valido(nome) else SEM_INFO,
        autor=autor,
        preco=preco,
        tecnica=tecnica,
        tecnica_original=tecnica_original,
        dimensoes=dimensoes,
        ano=ano,
        descricao=sem_info(descricao[:1200]),
        link=card.link,
        imagem=imagem if imagem_valida(imagem) else SEM_INFO,
    )


def obra_valida(obra: Obra) -> bool:
    if not link_lote_valido(obra.link):
        return False
    if not titulo_valido(obra.nome):
        return False
    # Não exige imagem, pois alguns lotes podem aparecer sem imagem real.
    return True


def extrair_dados_obra(detail_page: Page, card: CardInfo, delay: float = 0.25) -> Obra:
    html_detalhe = ""
    base_url = card.link

    try:
        detail_page.goto(card.link, wait_until="domcontentloaded", timeout=35_000)
        detail_page.wait_for_timeout(700)
        aceitar_cookies_e_popups(detail_page)
        html_detalhe = detail_page.content()
        base_url = detail_page.url or card.link
    except PlaywrightTimeoutError:
        print(f"  Timeout ao abrir detalhe: {card.link}", file=sys.stderr)
    except Exception as exc:
        print(f"  Erro ao abrir detalhe {card.link}: {exc}", file=sys.stderr)

    if delay:
        time.sleep(delay)

    return montar_obra_a_partir_das_fontes(card, html_detalhe, base_url)


def obra_to_row(obra: Obra) -> dict[str, Any]:
    return {
        "Nome da obra": obra.nome,
        "Autor": obra.autor,
        "Preço": obra.preco,
        "Técnica": obra.tecnica,
        "Técnica original": obra.tecnica_original,
        "Dimensões": obra.dimensoes,
        "Ano da obra": obra.ano,
        "Descrição": obra.descricao,
        "Link da obra": obra.link,
        "Link da imagem da obra": obra.imagem,
    }


def salvar_excel(df: pd.DataFrame, output: str) -> None:
    for col in EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = SEM_INFO
    df = df.reindex(columns=EXCEL_COLUMNS)

    pasta = os.path.dirname(os.path.abspath(output))
    if pasta:
        os.makedirs(pasta, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")
        ws = writer.sheets["Dados"]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        if "Preço" in headers:
            col_preco = get_column_letter(headers["Preço"])
            for row in range(2, ws.max_row + 1):
                ws[f"{col_preco}{row}"].number_format = '#,##0.00'

        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=12)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 70)


def scrape_leiloesbr(
    start_url: str = START_URL,
    max_obras: int = 0,
    output: str = "artes_LeiloesBR.xlsx",
    headless: bool = True,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    obras_salvas = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
        )
        page = context.new_page()
        detail_page = context.new_page()

        cards = descobrir_cards_listagem(page, start_url, max_obras=max_obras)
        print(f"Total de links únicos encontrados: {len(cards)}")

        for idx, card in enumerate(cards, start=1):
            print(f"Processando obra {idx}/{len(cards)}")
            try:
                obra = extrair_dados_obra(detail_page, card)
                if obra_valida(obra):
                    rows.append(obra_to_row(obra))
                    obras_salvas += 1
                    print(f"  OK: {obra.nome} | {obra.preco}")
                else:
                    print(f"  Ignorado: registro inválido ou genérico ({card.link})")
            except Exception as exc:
                print(f"  ERRO em {card.link}: {exc}", file=sys.stderr)

        context.close()
        browser.close()

    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    if not df.empty:
        df = df.drop_duplicates(subset=["Link da obra"], keep="first")

    salvar_excel(df, output)
    print(f"Obras salvas: {len(df)}")
    print(f"Arquivo gerado: {output}")
    return df


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scraper LeilõesBR sem miniaturas",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--max-obras", type=int, default=0, help="Quantidade máxima de obras; 0 tenta coletar o máximo possível.")
    parser.add_argument("--output", default="artes_LeiloesBR.xlsx", help="Arquivo Excel de saída.")
    parser.add_argument("--headless", default="true", help="true para navegador invisível; false para visualizar.")
    parser.add_argument("--url", default=START_URL, help="URL inicial da busca LeilõesBR.")
    args = parser.parse_args()

    data_execucao = datetime.now().strftime("%d/%m/%Y %H:%M")
    print(f"Iniciando LeilõesBR em {data_execucao}")
    scrape_leiloesbr(
        start_url=args.url,
        max_obras=args.max_obras,
        output=args.output,
        headless=str_to_bool(args.headless),
    )


if __name__ == "__main__":
    main()
