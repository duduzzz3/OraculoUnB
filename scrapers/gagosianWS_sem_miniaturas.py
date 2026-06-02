# importações
from playwright.sync_api import sync_playwright
import pyautogui
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
import requests
import os
import re


BASE_URL = "https://gagosianshop.com"
COLECAO_URL = f"{BASE_URL}/collections/prints-editions"


# pega texto de um locator com segurança

def texto_locator(locator):
    try:
        if locator.count() > 0:
            texto = locator.first.inner_text().strip()
            return texto if texto else "Sem informação"
    except Exception:
        pass
    return "Sem informação"


# pega o texto após os dois pontos

def extrair_apos_dois_pontos(texto):
    if not texto:
        return "Sem informação"
    texto = str(texto).strip()
    if ":" in texto:
        return texto.split(":", 1)[1].strip()
    return texto if texto else "Sem informação"


# pega cotação do dólar para converter preços em reais

def pegar_cotacao_dolar():
    try:
        resp = requests.get("https://economia.awesomeapi.com.br/json/last/USD-BRL", timeout=10)
        resp.raise_for_status()
        return float(resp.json()["USDBRL"]["bid"])
    except Exception:
        return 5.00


# limpa preço em texto e devolve só número

def limpar_preco_dolar(texto):
    if not texto or not str(texto).strip():
        return None

    numeros = re.findall(r"[\d,.]+", str(texto).strip())
    if not numeros:
        return None

    try:
        return float("".join(numeros).replace(",", "").strip())
    except ValueError:
        return None


# converte dólar em real

def converter_usd_para_brl(valor_usd, cotacao_dolar):
    if valor_usd is None:
        return "Sem informação"
    return round(valor_usd * cotacao_dolar, 2)


# normaliza urls de imagem

def normalizar_url_imagem(url):
    if not url or url == "Sem informação":
        return "Sem informação"
    url = str(url).strip()
    if url.startswith("//"):
        return "https:" + url
    if url.startswith("/"):
        return BASE_URL + url
    return url


# tenta pegar a melhor url da imagem no card

def extrair_melhor_url_imagem(locator_base):
    try:
        img = locator_base.locator("img").first
        candidatos = [
            img.get_attribute("src"),
            img.get_attribute("data-src"),
            img.get_attribute("data-original-src"),
            img.get_attribute("data-image"),
        ]

        srcset = img.get_attribute("srcset")
        if srcset:
            partes = [p.strip().split(" ")[0] for p in srcset.split(",") if p.strip()]
            if partes:
                candidatos.append(partes[-1])

        for candidato in candidatos:
            candidato = normalizar_url_imagem(candidato)
            if candidato and candidato != "Sem informação":
                return candidato
    except Exception:
        pass

    return "Sem informação"


# tenta pegar a imagem da página interna

def extrair_imagem_pagina_interna(detail_page):
    seletores = [
        'meta[property="og:image"]',
        "img.product-featured-img",
        "div.product-single__photo img",
        "img[data-product-featured-image]",
        "img",
    ]

    for seletor in seletores:
        try:
            loc = detail_page.locator(seletor)
            if loc.count() == 0:
                continue

            if seletor.startswith("meta"):
                url = loc.first.get_attribute("content")
            else:
                url = loc.first.get_attribute("src") or loc.first.get_attribute("data-src") or loc.first.get_attribute("data-original-src")
                if not url:
                    srcset = loc.first.get_attribute("srcset")
                    if srcset:
                        partes = [p.strip().split(" ")[0] for p in srcset.split(",") if p.strip()]
                        if partes:
                            url = partes[-1]

            url = normalizar_url_imagem(url)
            if url and url != "Sem informação":
                return url
        except Exception:
            continue

    return "Sem informação"


# limpa espaços duplicados

def limpar_espacos(texto):
    if not texto:
        return ""
    texto = str(texto).replace("\xa0", " ").replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", texto).strip()


# pega o bloco de especificações da obra

def extrair_bloco_especificacoes(detail_page):
    seletores = [
        "div.product-info__specifications-inner",
        "div.product-info__specifications",
        "div.product-info",
    ]

    for seletor in seletores:
        try:
            loc = detail_page.locator(seletor)
            if loc.count() > 0:
                texto = limpar_espacos(loc.first.inner_text())
                if texto:
                    return texto
        except Exception:
            continue

    return "Sem informação"


# extrai só o trecho das dimensões do bloco

def extrair_trecho_dimensoes(texto_specs):
    if not texto_specs or texto_specs == "Sem informação":
        return "Sem informação"

    match = re.search(
        r"Dimensions:\s*(.*?)(?=\bEdition:|\bSigned:|\bPrinter:|\bPublisher:|\bFramed:|\bArtist:|\bDate:|\bMedium:|$)",
        limpar_espacos(texto_specs),
        flags=re.IGNORECASE,
    )
    return limpar_espacos(match.group(1)) if match else "Sem informação"


# formata números de dimensões

def formatar_numero_dimensao(valor):
    try:
        numero = float(str(valor).replace(",", "."))
        return str(int(numero)) if numero.is_integer() else str(numero).rstrip("0").rstrip(".")
    except Exception:
        return str(valor).replace(",", ".").strip()


# padroniza dimensões no formato A x B cm

def extrair_dimensoes(texto_specs):
    trecho = extrair_trecho_dimensoes(texto_specs)
    if trecho == "Sem informação":
        return "Sem informação"

    match_print = re.search(r"Print:.*?\(([^()]*)\s*cm\)", trecho, flags=re.IGNORECASE)
    if match_print:
        numeros = re.findall(r"\d+(?:[.,]\d+)?", match_print.group(1))
        if len(numeros) >= 2:
            return f"{formatar_numero_dimensao(numeros[0])} x {formatar_numero_dimensao(numeros[1])} cm"

    grupos_cm = re.findall(r"\(([^()]*)\s*cm\)", trecho, flags=re.IGNORECASE)
    for grupo in grupos_cm:
        numeros = re.findall(r"\d+(?:[.,]\d+)?", grupo)
        if len(numeros) >= 2:
            return f"{formatar_numero_dimensao(numeros[0])} x {formatar_numero_dimensao(numeros[1])} cm"

    match_solto = re.search(r"(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)\s*cm", trecho, flags=re.IGNORECASE)
    if match_solto:
        return f"{formatar_numero_dimensao(match_solto.group(1))} x {formatar_numero_dimensao(match_solto.group(2))} cm"

    return "Sem informação"


# exporta dataframe para excel com miniaturas

def exportar_excel(df, arquivo_tmp, arquivo_final, mensagem):
    ordem = [
        "Título", "Autor", "Ano", "Técnica", "Dimensões",
        "Preço", "Descrição", "Link da obra", "Link da imagem da obra"
    ]

    for col in ordem:
        if col not in df.columns:
            df[col] = ""

    df = df.reindex(columns=ordem)

    with pd.ExcelWriter(arquivo_tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1, sheet_name="Dados")
        ws = writer.sheets["Dados"]

        ws["A1"] = mensagem

        header_row = 2
        data_start = 3
        headers_dict = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}

        if "Preço" in headers_dict:
            col_preco = get_column_letter(headers_dict["Preço"])
            for r in range(data_start, ws.max_row + 1):
                ws[f"{col_preco}{r}"].number_format = "0.00"

        ultima_col = get_column_letter(ws.max_column)
        ws.merge_cells(f"A1:{ultima_col}1")

    if os.path.exists(arquivo_final):
        os.remove(arquivo_final)
    os.rename(arquivo_tmp, arquivo_final)


with sync_playwright() as playwright:
    # abre navegador e páginas principal/detalhe
    cotacao_dolar = pegar_cotacao_dolar()
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    detail_page = context.new_page()

    # abre listagem inicial
    page.goto(COLECAO_URL, wait_until="domcontentloaded")

    dicionario_paginas = {}
    count = 1
    max_links = 150
    links_vistos = set()
    count_links = 0
    parar_tudo = False

    # percorre as páginas/carregamentos da listagem
    while True:
        artes = page.locator('.grid__item.wide--one-third.large--one-third.medium-down--one-half.grid--width-1\\/3')
        total = artes.count()
        print(f"{count}ª Página -> Obras encontradas: {total}")

        if total == 0:
            print("Nenhuma obra encontrada nessa página. Encerrando.")
            break

        lista_dicionarios = []

        # percorre os cards da página
        for i in range(total):
            arte = artes.nth(i)

            href = arte.locator("a.grid-link__image-container").first.get_attribute("href")
            link_arte = f"{BASE_URL}{href}" if href and href.startswith("/") else (href or "Sem informação")

            if link_arte in links_vistos:
                continue

            links_vistos.add(link_arte)
            count_links += 1

            if count_links > max_links:
                parar_tudo = True
                break

            titulo_arte = extrair_apos_dois_pontos(texto_locator(arte.locator("p.grid-link__title a")))
            link_imagem = extrair_melhor_url_imagem(arte)
            preco = converter_usd_para_brl(limpar_preco_dolar(texto_locator(arte.locator("p.heading-5"))), cotacao_dolar)

            dicionario = {
                "Título": titulo_arte,
                "Autor": "Sem informação",
                "Ano": "Sem informação",
                "Técnica": "Sem informação",
                "Dimensões": "Sem informação",
                "Preço": preco,
                "Descrição": "Sem informação",
                "Link da obra": link_arte,
                "Link da imagem da obra": link_imagem,
            }

            # abre página interna para buscar os dados restantes
            if link_arte != "Sem informação":
                try:
                    detail_page.goto(link_arte, wait_until="domcontentloaded")
                    detail_page.wait_for_timeout(1500)

                    texto_specs = extrair_bloco_especificacoes(detail_page)
                    dicionario["Descrição"] = texto_locator(detail_page.locator("div.product-info__description-inner"))
                    dicionario["Autor"] = extrair_apos_dois_pontos(texto_locator(detail_page.locator('p:has-text("Artist:")')))
                    dicionario["Técnica"] = extrair_apos_dois_pontos(texto_locator(detail_page.locator('p:has-text("Medium:")')))
                    dicionario["Ano"] = extrair_apos_dois_pontos(texto_locator(detail_page.locator('p:has-text("Date:")')))
                    dicionario["Dimensões"] = extrair_dimensoes(texto_specs)

                    if dicionario["Link da imagem da obra"] == "Sem informação":
                        dicionario["Link da imagem da obra"] = extrair_imagem_pagina_interna(detail_page)
                except Exception as e:
                    print(f"Erro ao abrir detalhes da obra: {link_arte} -> {e}")

            lista_dicionarios.append(dicionario)

            if count_links == max_links:
                print(f"Parando: cheguei em {max_links} links.")
                parar_tudo = True
                break

        dicionario_paginas[f"{count}ª Página"] = lista_dicionarios

        if parar_tudo:
            break

        botao_load_more = page.locator('button:has-text("Load More")')
        if botao_load_more.count() == 0:
            print("Chegamos ao fim!")
            break

        total_antes = artes.count()
        botao_load_more.first.click()
        page.wait_for_timeout(3000)
        total_depois = page.locator('.grid__item.wide--one-third.large--one-third.medium-down--one-half.grid--width-1\\/3').count()

        if total_depois <= total_antes:
            print("Não carregou novas obras. Encerrando.")
            break

        count += 1

    pyautogui.alert("Sistema Executado")
    context.close()
    browser.close()


# junta todas as páginas em uma lista única
lista_unificada = []
for lista in dicionario_paginas.values():
    lista_unificada.extend(lista)

df = pd.DataFrame(lista_unificada)
if not df.empty and "Link da obra" in df.columns:
    df = df.drop_duplicates(subset=["Link da obra"], keep="first")

site = "Gagosian"
data_execucao = datetime.now().strftime("%d/%m/%Y")
mensagem = f"Planilha alimentada com dados do site ({site}) no dia {data_execucao} | Cotação do dólar utilizada: {cotacao_dolar:.4f}"

exportar_excel(df, "artes_Gagosian_tmp.xlsx", "artes_Gagosian.xlsx", mensagem)
