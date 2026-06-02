# -*- coding: utf-8 -*-
"""
Scraper Saatchi Art - versão melhorada

Foco desta versão:
- Extrair dados principalmente da página interna da obra, não só do card da listagem.
- Padronizar dimensões no formato Altura x Largura x Profundidade cm.
- Converter dimensões em polegadas para centímetros.
- Extrair preço em USD e converter para BRL usando cotação online.
- Classificar técnica apenas dentro do rol definido.
- Gerar Excel com formatação consistente de cabeçalho, larguras, filtros, links e moedas.

Instalação:
    pip install playwright pandas openpyxl requests
    python -m playwright install chromium

Uso básico:
    python saatchiWS_melhorado.py --max 50

Uso com navegador invisível:
    python saatchiWS_melhorado.py --max 200 --headless true
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import Browser, BrowserContext, Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


BASE_URL = "https://www.saatchiart.com"
URL_LISTAGEM_PADRAO = "https://www.saatchiart.com/paintings"
ARQUIVO_EXCEL_PADRAO = "artes_SAATCHIART.xlsx"
ARQUIVO_TMP_PADRAO = "artes_SAATCHIART_tmp.xlsx"
DOLAR_PADRAO_FALLBACK = 5.00
SEM_INFO = "Sem informação"

# Desligado por padrão para manter o arquivo leve e preservar a lógica do script original "sem miniaturas".
# Se quiser miniaturas embutidas no Excel, troque para True.
INSERIR_MINIATURAS = False
PASTA_MINIATURAS = Path("_saatchi_miniaturas")

ROL_TECNICAS = [
    "Pintura acrílica",
    "Pintura a óleo",
    "Aquarela",
    "Guache",
    "Técnica mista",
    "Colagem",
    "Serigrafia",
    "Litografia",
    "Gravura",
    "Impressão fine art",
    "Giclée",
    "Fotografia",
    "Arte digital",
    "Escultura",
    "Cerâmica",
    "Metal",
    "Madeira",
    "Resina",
    "Têxtil",
    "Outros",
    "Sem informação",
]

ORDEM_COLUNAS = [
    "Nome da obra",
    "Autor",
    "Ano",
    "Técnica",
    "Técnica original",
    "Dimensões",
    "Altura (cm)",
    "Largura (cm)",
    "Profundidade (cm)",
    "Preço USD",
    "Preço BRL",
    "Cotação USD-BRL",
    "Descrição",
    "Link da obra",
    "Link da imagem da obra",
    "Status extração",
]

# Labels que aparecem no detalhe da obra e indicam início de outra seção.
LABELS_DETALHE = {
    "year created", "ano", "subject", "styles", "mediums", "rarity", "size", "ready to hang",
    "frame", "authenticity", "packaging", "delivery cost", "delivery time", "returns", "handling",
    "ships from", "customs", "about the artwork", "details and dimensions", "shipping and returns",
    "artist recognition", "view profile", "why saatchi art", "related searches", "top categories",
}

TIPOS_ARTE_INGLES = (
    "Painting", "Photography", "Photograph", "Sculpture", "Drawing", "Print", "Printmaking",
    "Mixed Media", "Collage", "Digital", "Ceramics", "Installation", "Textile", "New Media",
)


@dataclass
class DimensoesPadronizadas:
    texto: str = SEM_INFO
    altura_cm: float | None = None
    largura_cm: float | None = None
    profundidade_cm: float | None = None


@dataclass
class ObraSaatchi:
    nome_obra: str = SEM_INFO
    autor: str = SEM_INFO
    ano: str = SEM_INFO
    tecnica: str = SEM_INFO
    tecnica_original: str = SEM_INFO
    dimensoes: str = SEM_INFO
    altura_cm: float | None = None
    largura_cm: float | None = None
    profundidade_cm: float | None = None
    preco_usd: float | None = None
    preco_brl: float | None = None
    cotacao_usd_brl: float | None = None
    descricao: str = SEM_INFO
    link_obra: str = SEM_INFO
    link_imagem: str = SEM_INFO
    status_extracao: str = "OK"

    def para_linha_excel(self) -> dict[str, Any]:
        return {
            "Nome da obra": self.nome_obra,
            "Autor": self.autor,
            "Ano": self.ano,
            "Técnica": self.tecnica,
            "Técnica original": self.tecnica_original,
            "Dimensões": self.dimensoes,
            "Altura (cm)": self.altura_cm,
            "Largura (cm)": self.largura_cm,
            "Profundidade (cm)": self.profundidade_cm,
            "Preço USD": self.preco_usd,
            "Preço BRL": self.preco_brl,
            "Cotação USD-BRL": self.cotacao_usd_brl,
            "Descrição": self.descricao,
            "Link da obra": self.link_obra,
            "Link da imagem da obra": self.link_imagem,
            "Status extração": self.status_extracao,
        }


def configurar_log() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


def normalizar_espacos(texto: Any) -> str:
    if texto is None:
        return ""
    texto = str(texto).replace("\xa0", " ").replace("\u200b", " ")
    texto = re.sub(r"[ \t\r\f\v]+", " ", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


def limpar_linha(texto: Any) -> str:
    return re.sub(r"\s+", " ", normalizar_espacos(texto)).strip(" -–—|,;:")


def valor_textual(valor: Any) -> str:
    txt = limpar_linha(valor)
    return txt if txt else SEM_INFO


def abs_url(url: str | None, base: str = BASE_URL) -> str:
    if not url:
        return ""
    url = url.strip()
    if url.startswith("//"):
        return "https:" + url
    return urljoin(base, url)


def url_limpa(url: str) -> str:
    if not url:
        return ""
    partes = urlparse(url)
    return urlunparse((partes.scheme, partes.netloc, partes.path, "", "", ""))


def parse_float_br_en(valor: str) -> float | None:
    if not valor:
        return None
    texto = str(valor).strip()
    texto = re.sub(r"[^\d,.-]", "", texto)
    if not texto:
        return None

    # Ex.: 1,234.56 -> 1234.56 | 1.234,56 -> 1234.56 | 123,45 -> 123.45
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        numero = float(texto)
        if math.isnan(numero):
            return None
        return numero
    except ValueError:
        return None


def limpar_preco_usd(texto: str) -> float | None:
    if not texto:
        return None
    texto = normalizar_espacos(texto)
    match = re.search(r"US\$\s*([\d,.]+)|\$\s*([\d,.]+)", texto, flags=re.I)
    if not match:
        return None

    numero = match.group(1) or match.group(2)
    # Em USD, vírgula quase sempre é separador de milhar: $1,250 -> 1250.
    # Ainda preserva casos decimais incomuns: $125,50 -> 125.50.
    if re.fullmatch(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?", numero):
        numero = numero.replace(",", "")
    return parse_float_br_en(numero)


def pegar_cotacao_dolar() -> float:
    """Busca USD-BRL. Se a API falhar, usa fallback para o script não quebrar."""
    urls = [
        "https://economia.awesomeapi.com.br/json/last/USD-BRL",
        "https://economia.awesomeapi.com.br/last/USD-BRL",
    ]
    for url in urls:
        try:
            resp = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            data = resp.json()
            bid = data.get("USDBRL", {}).get("bid")
            cotacao = parse_float_br_en(str(bid))
            if cotacao and cotacao > 0:
                logging.info("Cotação USD-BRL obtida: %.4f", cotacao)
                return cotacao
        except Exception as exc:
            logging.warning("Falha ao buscar cotação em %s: %s", url, exc)
    logging.warning("Usando cotação fallback USD-BRL: %.2f", DOLAR_PADRAO_FALLBACK)
    return DOLAR_PADRAO_FALLBACK


def formatar_numero_cm(valor: float | None) -> str:
    if valor is None:
        return ""
    arredondado = round(float(valor), 2)
    txt = f"{arredondado:.2f}".rstrip("0").rstrip(".")
    return txt.replace(".", ",")


def padronizar_dimensoes(texto: str) -> DimensoesPadronizadas:
    """
    Converte textos como:
      Size: 19.7 W x 19.7 H x 0.8 D in
      50 W x 60 H x 2 D cm
      19.7 x 19.7 Inch

    Saída textual: Altura x Largura x Profundidade cm.
    Regra Saatchi: quando há W/H/D, W=largura, H=altura, D=profundidade.
    Quando não há rótulo, assume ordem Saatchi: W x H x D.
    """
    if not texto or not normalizar_espacos(texto):
        return DimensoesPadronizadas()

    original = normalizar_espacos(texto)
    texto_work = original.replace("×", "x").replace("X", "x")

    unidade = "cm"
    if re.search(r"\b(in|inch|inches)\b", texto_work, flags=re.I):
        unidade = "in"
    elif re.search(r"\bcm\b|centimeter|centimetre", texto_work, flags=re.I):
        unidade = "cm"

    partes = [p.strip() for p in re.split(r"\s+x\s+|\s+x\s*|\s*×\s*", texto_work) if p.strip()]
    medidas: dict[str, float] = {}
    sem_rotulo: list[float] = []

    for parte in partes:
        numero_match = re.search(r"\d+(?:[.,]\d+)?", parte)
        if not numero_match:
            continue
        valor = parse_float_br_en(numero_match.group(0))
        if valor is None:
            continue

        rotulo = None
        if re.search(r"\b(W|Width|Largura)\b", parte, flags=re.I):
            rotulo = "largura"
        elif re.search(r"\b(H|Height|Altura)\b", parte, flags=re.I):
            rotulo = "altura"
        elif re.search(r"\b(D|Depth|Profundidade)\b", parte, flags=re.I):
            rotulo = "profundidade"

        if unidade == "in":
            valor *= 2.54

        if rotulo:
            medidas[rotulo] = round(valor, 2)
        else:
            sem_rotulo.append(round(valor, 2))

    # Fallback: captura números mesmo quando o split falha.
    if not medidas and not sem_rotulo:
        nums = [parse_float_br_en(n) for n in re.findall(r"\d+(?:[.,]\d+)?", texto_work)]
        nums = [n for n in nums if n is not None]
        if unidade == "in":
            nums = [round(n * 2.54, 2) for n in nums]
        sem_rotulo = nums[:3]

    if sem_rotulo:
        # Ordem padrão do Saatchi quando não há W/H/D: Width x Height x Depth.
        if len(sem_rotulo) >= 1 and "largura" not in medidas:
            medidas["largura"] = sem_rotulo[0]
        if len(sem_rotulo) >= 2 and "altura" not in medidas:
            medidas["altura"] = sem_rotulo[1]
        if len(sem_rotulo) >= 3 and "profundidade" not in medidas:
            medidas["profundidade"] = sem_rotulo[2]

    altura = medidas.get("altura")
    largura = medidas.get("largura")
    profundidade = medidas.get("profundidade")

    if altura is None and largura is None:
        return DimensoesPadronizadas()

    if altura is not None and largura is not None and profundidade is not None:
        texto_fmt = f"{formatar_numero_cm(altura)} x {formatar_numero_cm(largura)} x {formatar_numero_cm(profundidade)} cm"
    elif altura is not None and largura is not None:
        texto_fmt = f"{formatar_numero_cm(altura)} x {formatar_numero_cm(largura)} cm"
    else:
        texto_fmt = SEM_INFO

    return DimensoesPadronizadas(texto=texto_fmt, altura_cm=altura, largura_cm=largura, profundidade_cm=profundidade)


def categorizar_tecnica(tecnica_original: str) -> str:
    """Classifica a técnica dentro do rol fechado informado pelo usuário."""
    if not tecnica_original or tecnica_original == SEM_INFO:
        return "Sem informação"

    t = tecnica_original.lower()
    t = t.replace("é", "e").replace("ê", "e").replace("á", "a").replace("à", "a")
    t = t.replace("ó", "o").replace("ô", "o").replace("í", "i").replace("ú", "u").replace("ç", "c")

    regras: list[tuple[str, Iterable[str]]] = [
        ("Giclée", ["giclee", "giclée"]),
        ("Pintura acrílica", ["acrylic", "acrilica", "acrylic painting"]),
        ("Pintura a óleo", ["oil", "oleo", "oil painting"]),
        ("Aquarela", ["watercolor", "watercolour", "aquarela"]),
        ("Guache", ["gouache", "guache"]),
        ("Técnica mista", ["mixed media", "tecnica mista"]),
        ("Colagem", ["collage", "colagem"]),
        ("Serigrafia", ["screenprint", "screen print", "serigraph", "silkscreen", "serigrafia"]),
        ("Litografia", ["lithograph", "lithography", "litografia"]),
        ("Impressão fine art", ["fine art print", "fine art prints", "archival print", "print on paper"]),
        ("Fotografia", ["photography", "photograph", "photo", "fotografia"]),
        ("Arte digital", ["digital", "new media", "arte digital"]),
        ("Cerâmica", ["ceramic", "ceramics", "clay", "porcelain", "stoneware", "ceramica"]),
        ("Resina", ["resin", "resina"]),
        ("Metal", ["metal", "bronze", "steel", "iron", "aluminum", "aluminium", "copper", "brass"]),
        ("Madeira", ["wood", "wooden", "timber", "madeira"]),
        ("Têxtil", ["textile", "fabric", "fiber", "fibre", "tapestry", "tecido", "textil"]),
        ("Escultura", ["sculpture", "sculptural", "escultura"]),
        ("Gravura", ["engraving", "etching", "woodcut", "linocut", "monotype", "monoprint", "printmaking", "gravura"]),
    ]

    for categoria, termos in regras:
        if any(termo in t for termo in termos):
            return categoria

    # Caso venha apenas "Painting" sem material, não inventa óleo/acrílica.
    if any(x in t for x in ["painting", "drawing", "print"]):
        return "Outros"
    return "Outros"


def linhas_visiveis(texto: str) -> list[str]:
    return [limpar_linha(l) for l in normalizar_espacos(texto).splitlines() if limpar_linha(l)]


def eh_label(linha: str) -> bool:
    l = limpar_linha(linha).lower().strip("#:")
    return l in LABELS_DETALHE


def extrair_valor_apos_label(texto: str, labels: Iterable[str], max_linhas: int = 6) -> str:
    labels_norm = {l.lower().strip().rstrip(":") for l in labels}
    linhas = linhas_visiveis(texto)
    for i, linha in enumerate(linhas):
        atual = linha.lower().strip().rstrip(":")
        if atual in labels_norm:
            valores: list[str] = []
            for prox in linhas[i + 1 : i + 1 + max_linhas]:
                if eh_label(prox):
                    break
                if prox in {",", "* * *", "-"}:
                    continue
                valores.append(prox)
            if valores:
                return limpar_linha(" ".join(valores))
    return ""


def extrair_titulo(page: Page, texto: str, url: str) -> str:
    candidatos: list[str] = []

    for seletor in ["h1", "[data-type='product-title']", "[class*='artworkTitle']", "[class*='ArtworkTitle']"]:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                candidatos.append(loc.inner_text(timeout=2000))
        except Exception:
            pass

    try:
        candidatos.append(page.title())
    except Exception:
        pass

    # Fallback pelo slug da URL.
    m_slug = re.search(r"/art/[^-]+-(.+?)/\d+/\d+/(?:view|print)", url)
    if m_slug:
        candidatos.append(m_slug.group(1).replace("-", " "))

    for cand in candidatos:
        cand = limpar_linha(cand)
        if not cand:
            continue

        # Ex.: "Sleeping Naked Woman - Nude Female Figure" Painting
        m_aspas = re.search(r"[\"“”'](.+?)[\"“”']\s+(?:" + "|".join(TIPOS_ARTE_INGLES) + r")\b", cand, flags=re.I)
        if m_aspas:
            return valor_textual(m_aspas.group(1))

        # Ex.: Sleeping Naked Woman - Nude Female Figure Painting by Daria Gerasimova | Saatchi Art
        cand = re.sub(r"\s+\|\s+Saatchi Art.*$", "", cand, flags=re.I)
        cand = re.sub(r"\s+by\s+.+$", "", cand, flags=re.I)
        cand = re.sub(r"\s+(?:" + "|".join(TIPOS_ARTE_INGLES) + r")$", "", cand, flags=re.I)
        cand = cand.strip(' "“”')
        if cand and cand.lower() not in {"saatchi art", "fax"}:
            return valor_textual(cand)

    return SEM_INFO


def extrair_autor(page: Page, texto: str) -> str:
    candidatos: list[str] = []

    try:
        titulo_pagina = page.title()
        m = re.search(r"\sby\s(.+?)\s\|\sSaatchi Art", titulo_pagina, flags=re.I)
        if m:
            candidatos.append(m.group(1))
    except Exception:
        pass

    # O autor normalmente é o primeiro link textual após o h1.
    try:
        loc = page.locator("xpath=(//h1/following::a[normalize-space()][1])").first
        if loc.count() > 0:
            candidatos.append(loc.inner_text(timeout=2000))
    except Exception:
        pass

    for seletor in [
        "[data-type='profile-artist-name']",
        "[class*='artistName']",
        "[class*='ArtistName']",
        "a[href*='/account/profile/']",
        "a[href*='/artist/']",
    ]:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                candidatos.append(loc.inner_text(timeout=2000))
        except Exception:
            pass

    linhas = linhas_visiveis(texto)
    for i, linha in enumerate(linhas[:80]):
        if re.match(r'^["“].+["”]\s+(?:' + "|".join(TIPOS_ARTE_INGLES) + r')\b', linha, flags=re.I):
            for prox in linhas[i + 1 : i + 5]:
                if prox and not re.search(r"\$|size:|\bview\b|\blog\s*in\b|register", prox, flags=re.I):
                    candidatos.append(prox)
                    break

    ignorar = {"log in", "register", "paintings", "photography", "sculpture", "drawings", "prints", "view profile"}
    for cand in candidatos:
        cand = limpar_linha(cand)
        if not cand:
            continue
        if cand.lower() in ignorar:
            continue
        if len(cand) > 80:
            continue
        if re.search(r"\$|size:|shipping|returns|saatchi", cand, flags=re.I):
            continue
        return valor_textual(cand)

    return SEM_INFO


def extrair_ano(texto: str) -> str:
    valor = extrair_valor_apos_label(texto, ["Year Created", "Ano", "Year"], max_linhas=3)
    m = re.search(r"\b(19\d{2}|20\d{2})\b", valor or texto)
    return m.group(1) if m else SEM_INFO


def extrair_preco_usd(page: Page, texto: str) -> float | None:
    seletores = [
        "text=/^\\$\\s*[\\d,]+(?:\\.\\d{2})?$/",
        "h4:has-text('$')",
        "h5:has-text('$')",
        "h6:has-text('$')",
        "[class*='price' i]",
        "[data-type*='price' i]",
    ]
    for seletor in seletores:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                preco = limpar_preco_usd(loc.inner_text(timeout=2000))
                if preco is not None:
                    return preco
        except Exception:
            pass

    # Evita capturar números de dimensão: só captura com símbolo de dólar.
    precos = [limpar_preco_usd(m.group(0)) for m in re.finditer(r"(?:US\$|\$)\s*[\d,.]+", texto, flags=re.I)]
    precos = [p for p in precos if p is not None]
    return precos[0] if precos else None


def texto_primeiro_locator_visivel(page: Page, seletores: Iterable[str], timeout_ms: int = 2000) -> str:
    """Retorna o primeiro texto visível encontrado nos seletores informados."""
    for seletor in seletores:
        try:
            loc = page.locator(seletor)
            total = loc.count()
            for i in range(min(total, 8)):
                item = loc.nth(i)
                if not item.is_visible(timeout=800):
                    continue
                txt = limpar_linha(item.inner_text(timeout=timeout_ms))
                if txt:
                    return txt
        except Exception:
            continue
    return ""


def parece_tecnica_principal(texto: str) -> bool:
    """Valida se o texto parece a técnica/mídia da obra, e não um item de menu ou recomendação."""
    if not texto:
        return False
    linha = limpar_linha(texto)
    if len(linha) > 120:
        return False

    rejeitar = [
        "All Artworks", "Paintings /", "More From", "Find Similar", "Shipping", "Returns",
        "Trustpilot", "Add To Cart", "Make An Offer", "People Are Interested", "Gift Guide",
    ]
    if any(x.lower() in linha.lower() for x in rejeitar):
        return False

    # A linha principal do Saatchi costuma ser exatamente assim:
    # "Painting, Acrylic on Canvas", "Photography, Digital on Paper", etc.
    padrao_tipo = re.compile(r"^(?:" + "|".join(map(re.escape, TIPOS_ARTE_INGLES)) + r")\s*,\s*.+", flags=re.I)
    if padrao_tipo.search(linha):
        return True

    # Fallback: aceita materiais/técnicas claros quando vierem sem o tipo inicial.
    termos_midia = [
        "acrylic", "oil", "watercolor", "watercolour", "gouache", "mixed media", "collage",
        "screenprint", "serigraph", "silkscreen", "lithograph", "giclee", "giclée",
        "fine art print", "archival print", "digital", "bronze", "steel", "wood", "resin",
        "ceramic", "textile", "canvas", "paper", "metal",
    ]
    return any(t in linha.lower() for t in termos_midia)


def limpar_tecnica_original_extraida(texto: str) -> str:
    """Remove sobras comuns e devolve só a técnica/mídia."""
    txt = limpar_linha(texto)
    txt = re.sub(r"^Mediums?\s*:?\s*", "", txt, flags=re.I)
    txt = re.sub(r"^M[ií]dia\s*:?\s*", "", txt, flags=re.I)
    txt = re.sub(r"\s*,\s*", ", ", txt)
    txt = re.sub(r"\s{2,}", " ", txt).strip(" -–—|,;:")
    return valor_textual(txt)


def extrair_tecnica_original(page: Page, texto: str) -> str:
    """
    Extrai a técnica do bloco destacado na página interna:
        div[class*='PdpImageAndInfo_artworkInfoBox'] > p:first-of-type

    Esse caminho é o mais confiável porque fica logo abaixo do autor e acima das dimensões.
    O fallback por texto completo só roda se esse bloco não estiver disponível.
    """
    seletores_prioritarios = [
        "div[class*='PdpImageAndInfo_artworkInfoBox'] > p:first-child",
        "div[class*='PdpImageAndInfo_artworkInfoBox'] > p:nth-child(1)",
        "xpath=(//h1/following::div[contains(@class,'PdpImageAndInfo_artworkInfoBox')][1]/p[1])",
    ]

    tecnica_direta = texto_primeiro_locator_visivel(page, seletores_prioritarios)
    if parece_tecnica_principal(tecnica_direta):
        return limpar_tecnica_original_extraida(tecnica_direta)

    # Fallback controlado: procura somente linhas completas do padrão "Tipo, Técnica on Suporte".
    linhas = linhas_visiveis(texto)
    padrao_linha_principal = re.compile(
        r"^(?:" + "|".join(map(re.escape, TIPOS_ARTE_INGLES)) + r")\s*,\s*[^$]{2,120}$",
        flags=re.I,
    )

    for linha in linhas[:180]:
        if padrao_linha_principal.search(linha) and parece_tecnica_principal(linha):
            return limpar_tecnica_original_extraida(linha)

    # Último fallback: bloco Mediums. Aqui pode vir "Acrylic, Canvas" em vez de
    # "Painting, Acrylic on Canvas", mas é melhor do que classificar errado.
    valor_mediums = extrair_valor_apos_label(texto, ["Mediums", "Medium", "Mídia", "Materiais"], max_linhas=8)
    if parece_tecnica_principal(valor_mediums):
        return limpar_tecnica_original_extraida(valor_mediums)

    return SEM_INFO

def extrair_dimensoes(texto: str) -> DimensoesPadronizadas:
    linhas = linhas_visiveis(texto)

    # Procura "Size:" seguido da medida na mesma linha ou na próxima.
    for i, linha in enumerate(linhas):
        if re.match(r"^Size\s*:?$", linha, flags=re.I):
            for prox in linhas[i + 1 : i + 4]:
                if re.search(r"\d+(?:[.,]\d+)?\s*(?:W|H|D|x|×|cm|in|inch|inches)", prox, flags=re.I):
                    dim = padronizar_dimensoes(prox)
                    if dim.texto != SEM_INFO:
                        return dim
        m = re.search(r"Size\s*:\s*(.+)$", linha, flags=re.I)
        if m:
            dim = padronizar_dimensoes(m.group(1))
            if dim.texto != SEM_INFO:
                return dim

    # Fallback global para padrões com W/H/D.
    m = re.search(
        r"\d+(?:[.,]\d+)?\s*W\s*[x×]\s*\d+(?:[.,]\d+)?\s*H(?:\s*[x×]\s*\d+(?:[.,]\d+)?\s*D)?\s*(?:cm|in|inch|inches)?",
        texto,
        flags=re.I,
    )
    if m:
        return padronizar_dimensoes(m.group(0))

    return DimensoesPadronizadas()


def extrair_descricao(page: Page, texto: str) -> str:
    for seletor in ["p[data-type='description']", "[data-type='description']", "[class*='description' i]"]:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                desc = limpar_linha(loc.inner_text(timeout=2000))
                if desc and len(desc) > 20:
                    return desc
        except Exception:
            pass

    m = re.search(
        r"ABOUT THE ARTWORK\s*(.*?)(?:READ MORE|Year Created\s*:|DETAILS AND DIMENSIONS|SHIPPING AND RETURNS)",
        texto,
        flags=re.I | re.S,
    )
    if m:
        desc = limpar_linha(m.group(1))
        if desc:
            return desc

    return SEM_INFO


def maior_url_srcset(srcset: str) -> str:
    if not srcset:
        return ""
    melhor_url = ""
    melhor_peso = -1
    for parte in srcset.split(","):
        pedacos = parte.strip().split()
        if not pedacos:
            continue
        url = pedacos[0]
        peso = 0
        if len(pedacos) > 1:
            m = re.search(r"(\d+)(?:w|x)", pedacos[1])
            if m:
                peso = int(m.group(1))
        if peso >= melhor_peso:
            melhor_url = url
            melhor_peso = peso
    return melhor_url


def extrair_imagem(page: Page, titulo: str = "") -> str:
    # Metatags primeiro, se existirem no HTML renderizado.
    for seletor in [
        "meta[property='og:image']",
        "meta[name='twitter:image']",
        "meta[property='og:image:secure_url']",
    ]:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                url = loc.get_attribute("content")
                if url:
                    return abs_url(url, page.url)
        except Exception:
            pass

    try:
        imgs = page.locator("img").evaluate_all(
            """
            (els) => els.map((img) => ({
                src: img.currentSrc || img.src || img.getAttribute('src') || '',
                dataSrc: img.getAttribute('data-src') || '',
                srcset: img.getAttribute('srcset') || '',
                alt: img.getAttribute('alt') || '',
                width: img.naturalWidth || img.width || 0,
                height: img.naturalHeight || img.height || 0
            }))
            """
        )
    except Exception:
        imgs = []

    titulo_low = (titulo or "").lower()
    candidatos: list[tuple[int, str]] = []
    for img in imgs:
        src = img.get("src") or img.get("dataSrc") or maior_url_srcset(img.get("srcset", ""))
        src = abs_url(src, page.url)
        if not src or src.startswith("data:"):
            continue

        alt = (img.get("alt") or "").lower()
        url_low = src.lower()
        if any(x in url_low for x in ["logo", "avatar", "profile", "curator", "background", "room-background"]):
            continue
        if any(x in alt for x in ["artist", "curator", "background", "view in a room background"]):
            continue

        score = 0
        if "saatchi" in url_low:
            score += 2
        if "original" in alt or "artwork" in alt:
            score += 4
        if titulo_low and titulo_low[:20] in alt:
            score += 5
        score += min(int(img.get("width") or 0), 2000) // 250
        score += min(int(img.get("height") or 0), 2000) // 250
        candidatos.append((score, src))

    if candidatos:
        candidatos.sort(reverse=True, key=lambda x: x[0])
        return candidatos[0][1]

    return SEM_INFO


def aceitar_cookies_e_popups(page: Page) -> None:
    seletores = [
        "button:has-text('Accept')",
        "button:has-text('Accept All')",
        "button:has-text('I Accept')",
        "button:has-text('Agree')",
        "button:has-text('Got it')",
        "button[aria-label='Close']",
        "button:has-text('Close')",
    ]
    for seletor in seletores:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0 and loc.is_visible(timeout=1000):
                loc.click(timeout=1500)
                page.wait_for_timeout(500)
        except Exception:
            pass


def rolar_para_carregar(page: Page, passos: int = 8, pausa_ms: int = 700) -> None:
    try:
        page.evaluate("window.scrollTo(0, 0)")
    except Exception:
        pass
    for _ in range(passos):
        try:
            page.mouse.wheel(0, 1500)
            page.wait_for_timeout(pausa_ms)
        except Exception:
            break


def eh_link_obra_saatchi(url: str) -> bool:
    if not url:
        return False
    url = url_limpa(abs_url(url))
    return bool(re.search(r"/art/[^/]+/\d+/\d+/(?:view|print)$", url))


def coletar_links_da_listagem(page: Page, limite: int) -> list[str]:
    aceitar_cookies_e_popups(page)
    try:
        page.wait_for_selector("a[href*='/art/']", timeout=20000)
    except PlaywrightTimeoutError:
        logging.warning("Nenhum link de obra apareceu na listagem dentro do timeout.")

    rolar_para_carregar(page)

    hrefs: list[str] = []
    try:
        hrefs = page.locator("a[href*='/art/']").evaluate_all("els => els.map(a => a.href || a.getAttribute('href') || '')")
    except Exception as exc:
        logging.warning("Falha ao ler links da listagem: %s", exc)

    links: list[str] = []
    vistos: set[str] = set()
    for href in hrefs:
        link = url_limpa(abs_url(href, page.url))
        if not eh_link_obra_saatchi(link):
            continue
        if link.endswith("/print"):
            link = link[:-6] + "/view"
        if link not in vistos:
            vistos.add(link)
            links.append(link)
        if len(links) >= limite:
            break

    return links


def proxima_url_por_parametro(url_atual: str, pagina: int) -> str:
    parsed = urlparse(url_atual)
    query = parse_qs(parsed.query)
    query["page"] = [str(pagina)]
    nova_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, nova_query, parsed.fragment))


def ir_para_proxima_pagina(page: Page, pagina_destino: int) -> bool:
    url_antes = page.url
    seletores = [
        "a[aria-label='Next page']",
        "button[aria-label='Next page']",
        "a[rel='next']",
        "a:has-text('Next')",
        "button:has-text('Next')",
    ]
    for seletor in seletores:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0 and loc.is_visible(timeout=1500):
                loc.click(timeout=3000)
                page.wait_for_load_state("domcontentloaded", timeout=30000)
                page.wait_for_timeout(2500)
                if page.url != url_antes:
                    return True
        except Exception:
            pass

    # Fallback por parâmetro ?page=N.
    try:
        proxima = proxima_url_por_parametro(url_antes, pagina_destino)
        page.goto(proxima, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(2500)
        return page.url != url_antes
    except Exception:
        return False


def extrair_obra(detail_page: Page, link: str, cotacao_dolar: float, tentativas: int = 2) -> ObraSaatchi:
    erros: list[str] = []
    for tentativa in range(1, tentativas + 1):
        try:
            detail_page.goto(link, wait_until="domcontentloaded", timeout=70000)
            try:
                detail_page.wait_for_load_state("networkidle", timeout=12000)
            except Exception:
                pass
            detail_page.wait_for_timeout(1500)
            aceitar_cookies_e_popups(detail_page)

            texto = detail_page.locator("body").inner_text(timeout=10000)
            texto = normalizar_espacos(texto)

            nome = extrair_titulo(detail_page, texto, link)
            autor = extrair_autor(detail_page, texto)
            ano = extrair_ano(texto)
            preco_usd = extrair_preco_usd(detail_page, texto)
            preco_brl = round(preco_usd * cotacao_dolar, 2) if preco_usd is not None else None
            tecnica_original = extrair_tecnica_original(detail_page, texto)
            tecnica = categorizar_tecnica(tecnica_original)
            dim = extrair_dimensoes(texto)
            descricao = extrair_descricao(detail_page, texto)
            imagem = extrair_imagem(detail_page, nome)

            status = []
            for campo, valor in {
                "nome": nome,
                "autor": autor,
                "ano": ano,
                "técnica": tecnica,
                "dimensões": dim.texto,
                "preço": preco_usd,
                "imagem": imagem,
            }.items():
                if valor in [None, "", SEM_INFO]:
                    status.append(f"sem {campo}")
            status_txt = "OK" if not status else "; ".join(status)

            return ObraSaatchi(
                nome_obra=nome,
                autor=autor,
                ano=ano,
                tecnica=tecnica,
                tecnica_original=tecnica_original,
                dimensoes=dim.texto,
                altura_cm=dim.altura_cm,
                largura_cm=dim.largura_cm,
                profundidade_cm=dim.profundidade_cm,
                preco_usd=preco_usd,
                preco_brl=preco_brl,
                cotacao_usd_brl=round(cotacao_dolar, 4),
                descricao=descricao,
                link_obra=link,
                link_imagem=imagem,
                status_extracao=status_txt,
            )
        except Exception as exc:
            erros.append(f"tentativa {tentativa}: {exc}")
            logging.warning("Falha ao extrair %s (%s/%s): %s", link, tentativa, tentativas, exc)
            detail_page.wait_for_timeout(1500)

    return ObraSaatchi(link_obra=link, status_extracao="ERRO: " + " | ".join(erros))


def baixar_imagem(url: str, destino: Path) -> Path | None:
    if not url or url == SEM_INFO:
        return None
    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        resp = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "").lower()
        if "image" not in content_type:
            return None
        destino.write_bytes(resp.content)
        return destino
    except Exception as exc:
        logging.warning("Falha ao baixar imagem %s: %s", url, exc)
        return None


def exportar_excel(df: pd.DataFrame, arquivo_tmp: str, arquivo_final: str, mensagem: str) -> None:
    for col in ORDEM_COLUNAS:
        if col not in df.columns:
            df[col] = None
    df = df.reindex(columns=ORDEM_COLUNAS)

    with pd.ExcelWriter(arquivo_tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=1, sheet_name="Dados")

    wb = load_workbook(arquivo_tmp)
    ws = wb["Dados"]

    ultima_col = get_column_letter(ws.max_column)
    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"] = mensagem
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    header_row = 2
    data_start = 3
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers_dict = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    larguras = {
        "Nome da obra": 36,
        "Autor": 26,
        "Ano": 10,
        "Técnica": 20,
        "Técnica original": 28,
        "Dimensões": 28,
        "Altura (cm)": 13,
        "Largura (cm)": 13,
        "Profundidade (cm)": 18,
        "Preço USD": 14,
        "Preço BRL": 16,
        "Cotação USD-BRL": 16,
        "Descrição": 70,
        "Link da obra": 45,
        "Link da imagem da obra": 55,
        "Status extração": 28,
    }

    for nome_col, largura in larguras.items():
        if nome_col in headers_dict:
            ws.column_dimensions[get_column_letter(headers_dict[nome_col])].width = largura

    colunas_numero = ["Altura (cm)", "Largura (cm)", "Profundidade (cm)", "Cotação USD-BRL"]
    colunas_moeda_usd = ["Preço USD"]
    colunas_moeda_brl = ["Preço BRL"]
    colunas_link = ["Link da obra", "Link da imagem da obra"]

    for r in range(data_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        for col in colunas_numero:
            if col in headers_dict:
                ws.cell(r, headers_dict[col]).number_format = "0.00"

        for col in colunas_moeda_usd:
            if col in headers_dict:
                ws.cell(r, headers_dict[col]).number_format = '[$$-en-US] #,##0.00'

        for col in colunas_moeda_brl:
            if col in headers_dict:
                ws.cell(r, headers_dict[col]).number_format = 'R$ #,##0.00'

        for col in colunas_link:
            if col in headers_dict:
                cell = ws.cell(r, headers_dict[col])
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        ws.row_dimensions[r].height = 60 if not INSERIR_MINIATURAS else 95

    # Congelar topo, aplicar autofiltro e melhorar visualização.
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ultima_col}{ws.max_row}"

    if INSERIR_MINIATURAS:
        try:
            from openpyxl.drawing.image import Image as XLImage

            if "Link da imagem da obra" in headers_dict:
                col_img_link = headers_dict["Link da imagem da obra"]
                col_miniatura = ws.max_column + 1
                ws.cell(header_row, col_miniatura, "Miniatura")
                ws.cell(header_row, col_miniatura).font = Font(bold=True, color="1F1F1F")
                ws.cell(header_row, col_miniatura).fill = header_fill
                ws.cell(header_row, col_miniatura).alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_miniatura)].width = 18

                for r in range(data_start, ws.max_row + 1):
                    url = ws.cell(r, col_img_link).value
                    if not url:
                        continue
                    caminho = baixar_imagem(str(url), PASTA_MINIATURAS / f"obra_{r}.jpg")
                    if not caminho:
                        continue
                    img = XLImage(str(caminho))
                    img.width = 90
                    img.height = 90
                    ws.add_image(img, f"{get_column_letter(col_miniatura)}{r}")
        except Exception as exc:
            logging.warning("Miniaturas não foram inseridas: %s", exc)

    arquivo_tmp_path = Path(arquivo_tmp)
    arquivo_final_path = Path(arquivo_final)
    if arquivo_final_path.exists():
        arquivo_final_path.unlink()
    wb.save(arquivo_tmp_path)
    arquivo_tmp_path.replace(arquivo_final_path)


def montar_dataframe(obras: list[ObraSaatchi]) -> pd.DataFrame:
    df = pd.DataFrame([obra.para_linha_excel() for obra in obras])
    if not df.empty and "Link da obra" in df.columns:
        df = df.drop_duplicates(subset=["Link da obra"], keep="first")
    return df


def executar_scraper(
    url_listagem: str,
    max_obras: int,
    max_paginas: int,
    headless: bool,
    arquivo_excel: str,
    slow_mo: int = 0,
) -> None:
    cotacao_dolar = pegar_cotacao_dolar()
    obras: list[ObraSaatchi] = []
    links_vistos: set[str] = set()

    with sync_playwright() as playwright:
        browser: Browser = playwright.chromium.launch(
            headless=headless,
            slow_mo=slow_mo,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox",
            ],
        )
        context: BrowserContext = browser.new_context(
            viewport={"width": 1440, "height": 1200},
            locale="en-US",
            timezone_id="America/Sao_Paulo",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        detail_page = context.new_page()

        try:
            logging.info("Abrindo listagem: %s", url_listagem)
            page.goto(url_listagem, wait_until="domcontentloaded", timeout=70000)
            page.wait_for_timeout(3000)

            pagina_atual = 1
            while len(obras) < max_obras and pagina_atual <= max_paginas:
                restante = max_obras - len(obras)
                links = coletar_links_da_listagem(page, limite=max(restante, 24))
                novos_links = [link for link in links if link not in links_vistos]
                logging.info("Página %s: %s links de obras encontrados, %s novos.", pagina_atual, len(links), len(novos_links))

                if not novos_links:
                    logging.warning("Nenhum link novo nesta página. Tentando próxima página.")

                for idx, link in enumerate(novos_links, start=1):
                    if len(obras) >= max_obras:
                        break
                    links_vistos.add(link)
                    logging.info("Extraindo %s/%s: %s", len(obras) + 1, max_obras, link)
                    obra = extrair_obra(detail_page, link, cotacao_dolar)
                    obras.append(obra)

                if len(obras) >= max_obras:
                    break

                pagina_atual += 1
                if not ir_para_proxima_pagina(page, pagina_destino=pagina_atual):
                    logging.info("Não foi possível avançar. Encerrando paginação.")
                    break

        finally:
            context.close()
            browser.close()

    df = montar_dataframe(obras)
    data_execucao = datetime.now().strftime("%d/%m/%Y %H:%M")
    mensagem = (
        f"Planilha alimentada com dados do site Saatchi Art em {data_execucao} | "
        f"Cotação USD-BRL utilizada: {cotacao_dolar:.4f} | "
        f"Total de obras: {len(df)}"
    )
    exportar_excel(df, ARQUIVO_TMP_PADRAO, arquivo_excel, mensagem)
    logging.info("Excel gerado: %s", arquivo_excel)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scraper Saatchi Art com extração robusta e Excel formatado.")
    parser.add_argument("--url", default=URL_LISTAGEM_PADRAO, help="URL inicial da listagem do Saatchi Art.")
    parser.add_argument("--max", type=int, default=50, help="Quantidade máxima de obras para extrair.")
    parser.add_argument("--paginas", type=int, default=20, help="Quantidade máxima de páginas da listagem para percorrer.")
    parser.add_argument("--headless", default="false", choices=["true", "false"], help="Rodar navegador invisível.")
    parser.add_argument("--saida", default=ARQUIVO_EXCEL_PADRAO, help="Nome do arquivo Excel final.")
    parser.add_argument("--slow-mo", type=int, default=0, help="Atraso em ms entre ações do Playwright, útil para depuração.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    configurar_log()
    args = parse_args(argv or sys.argv[1:])
    headless = args.headless.lower() == "true"

    executar_scraper(
        url_listagem=args.url,
        max_obras=max(1, args.max),
        max_paginas=max(1, args.paginas),
        headless=headless,
        arquivo_excel=args.saida,
        slow_mo=max(0, args.slow_mo),
    )


if __name__ == "__main__":
    main()
